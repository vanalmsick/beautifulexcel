# -*- coding: utf-8 -*-
import os
import warnings
from typing import Any, Union, List, Tuple #, Literal
import datetime
import numpy as np
import pandas as pd
import openpyxl
import yaml


from beautifulexcel.utils import (
    flatten_dict,
    deepen_dict,
    get_custom_styles,
    dict_extend_with_dict,
    is_valid_excel_cell,
)


class Sheet:
    """Base Excel Sheet Class which contains all the methods that can be applied to a sheet"""

    def __init__(self, excelwriter, sheet_name, use_theme_style, col_widths):
        self.excelwriter = excelwriter
        self.writer = self.excelwriter.writer
        self.sheet_name = sheet_name
        self.use_theme_style = use_theme_style
        self.col_widths = col_widths

        if use_theme_style:
            self.style_base = self._extend_style_args(self.excelwriter.theme.get("general", {}))
        else:
            self.style_base = {}

    def _extend_style_args(self, style_args):
        """Uses the theme presets and replaces them with the actual style kwargs"""
        self.preset_styles = preset_styles = self.excelwriter.theme.get("preset", {})

        # already extended dict style
        if type(style_args) is dict:
            return style_args
        # single preset name
        elif type(style_args) is str:
            style_i = preset_styles.get(style_args)
            if style_i is None:
                raise Exception(f'You are referencing the style preset "{style_i}" but it is not defined in the theme.')
            return preset_styles.get(style_args)
        # list of preset names or extended dict
        elif type(style_args) is list:
            _final_style = {}
            for i_style_args in style_args:
                # is preset name
                if type(i_style_args) is str:
                    if style_i is None:
                        style_i = preset_styles.get(i_style_args)
                        raise Exception(
                            f'You are referencing the style preset "{i_style_args}" but it is not defined in the theme.'
                        )
                    _final_style = {**_final_style, **style_i}
                # is already extended dict
                else:
                    _final_style = {**_final_style, **i_style_args}
            return _final_style

    def apply_cell_style(self, row_num, col_num, style):
        ws = self.ws
        style_deep = deepen_dict(style)
        cell = ws.cell(row=row_num, column=col_num)

        for style_type, kwargs in style_deep.items():
            if style_type.lower() == "font":
                cell.font = openpyxl.styles.Font(**kwargs)
            elif style_type.lower() == "numfmt" or style_type.lower() == "numberformat":
                cell.number_format = kwargs
            elif style_type.lower() == "align" or style_type.lower() == "alignment":
                cell.alignment = openpyxl.styles.Alignment(**kwargs)
            elif style_type.lower() == "fill" or style_type.lower() == "pfill" or style_type.lower() == "patternfill":
                if type(kwargs) is not dict:
                    kwargs = {"patternType": "solid", "fgColor": kwargs}
                cell.fill = openpyxl.styles.PatternFill(**kwargs)
            elif style_type.lower() == "gfill" or style_type.lower() == "gradientfill":
                cell.fill = openpyxl.styles.GradientFill(**kwargs)
            elif style_type.lower() == "border" or style_type.lower() == "borders":
                border_kwargs = {}
                for border_type, border_props in kwargs.items():
                    border_kwargs[border_type] = openpyxl.styles.Side(**border_props)
                cell.border = openpyxl.styles.Border(**border_kwargs)
            elif style_type.lower() == "protection":
                cell.font = openpyxl.styles.Protection(**kwargs)
            else:
                raise Exception(
                    f'Unknown style type "{style_type}". Available style types are: font, numberformat, align, fill, patternfill, gradientfill, borders, and protection.'
                )

        return cell

    def apply_range_style(self, ws, style):
        raise Exception("Not yet defined function .apply_range_style()")

    def change_col_widths(self, col_widths: dict):
        for col, width in col_widths.items():
            if type(col) is int:
                col = openpyxl.utils.get_column_letter(col + 1)
            self.ws.column_dimensions[col].width = width

    def add_data_validation(
            self,
            ref: Union[str, List[str]],
            type: str, #Literal["list", "whole", "decimal", "date", "time", "textLength", "formula"],
            props: Union[str, List[str], int, float, Tuple[int, float]] = None,
            operator: str = None, #Literal["between", "greaterThan", "greaterThanOrEqual", "equal", "notEqual", "lessThan", "lessThanOrEqual"] = None,
            **kwargs
        ):
        """
        Add data validation to worksheet

        Args:
            ref (str or list of str): Cell range reference e.g. "A1:C5" or ["A1:C5", "A21:Z55"]
            type (str): Data valiation type. Options: list, whole, decimal, date, time, textLength, formula
            operator (str): Validation mathematical operator: "between", "greaterThan", "greaterThanOrEqual", "equal", "notEqual", "lessThan", "lessThanOrEqual"
            props (any): Depending on vaildation type the properties: list = ["Option 1", "Option 2"], numeric = single number OR upper & lower boundary (1, 100), formla
            kwargs (dict): Any futher kwarg specified in [openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.datavalidation.html](https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.datavalidation.html#openpyxl.worksheet.datavalidation.DataValidation)

        Examples:
            - ```sheet1.add_data_validation(ref="A1:C5", type="whole")```
            - ```sheet1.add_data_validation(ref="A1:C5", type="decimal", operator="greaterThan", props=0)```
            - ```sheet1.add_data_validation(ref="A1:C5", type="textLength", props=10)```
            - ```sheet1.add_data_validation(ref="A1:C5", type="list", props=["Yes", "No"])```
            - ```sheet1.add_data_validation(ref="A1:C5", type="whole", operator="between", props=[0, 100])```
        """
        if isinstance(props, list) or isinstance(props, tuple):
            if type.lower() == 'list':
                formula1 = '"' + ','.join(props) + '"'
                formula2 = None
            else:
                formula1, formula2 = props
        else:
            formula1 = props
            formula2 = None
        dv = openpyxl.worksheet.datavalidation.DataValidation(type=type, formula1=formula1, formula2=formula2, operator=operator, **kwargs)
        self.ws.add_data_validation(dv)

        if isinstance(ref, str):
            ref = [ref]
        for i_ref in ref:
            if self.__class__.__name__ == 'DataframeSheet':
                i_ref = self.util_range_ref_from_coordinates(self.util_get_range_coordinates(i_ref, enrich_dimensions=True))
            if i_ref is not None:
                dv.add(i_ref)


    def merge_cells(self, ref: str):
        """
        Merge several cells

        Args:
            ref (str): Cell range reference e.g. 'A1:C5'

        Example:
            ```sheet1.merge_cells(ref='A1:C5')```
        """
        if isinstance(ref, str):
            ref = [ref]
        for i_ref in ref:
            if self.__class__.__name__ == 'DataframeSheet':
                i_ref = self.util_range_ref_from_coordinates(self.util_get_range_coordinates(i_ref, enrich_dimensions=True))
            if i_ref is not None:
                self.ws.merge_cells(i_ref)


    def __group_cols_rows(self, ref, axis, hidden=False, outline_level=1):
        """Group several rows or cols"""
        if isinstance(ref, str):
            ref = [ref]
            if isinstance(hidden, bool):
                hidden = [hidden] * len(ref)
            if isinstance(outline_level, int):
                outline_level = [outline_level] * len(ref)
        for i_ref, i_hidden, i_level in zip(ref, hidden, outline_level):
            if self.__class__.__name__ == 'DataframeSheet':
                i_ref = self.util_range_ref_from_coordinates(self.util_get_range_coordinates(i_ref))
            if i_ref is not None:
                i_ref_start, i_ref_end = i_ref.split(':')
                if axis == 'columns':
                    self.ws.column_dimensions.group(i_ref_start, i_ref_end, hidden=i_hidden, outline_level=i_level)
                elif axis == 'rows':
                    self.ws.row_dimensions.group(i_ref_start, i_ref_end, hidden=i_hidden, outline_level=i_level)


    def group_columns(self, ref: str):
        """
        Group several columns

        Args:
            ref (str): Excel column reference e.g. 'A:C'

        Example:
            ```sheet1.group_columns(ref='1:2')```
        """
        return self.__group_cols_rows(ref=ref, axis='columns')

    def group_rows(self, ref):
        """
        Group several rows

        Args:
            ref (str): Excel row reference e.g. '1:3'

        Example:
            ```sheet1.group_rows(ref='1:5')```
        """
        return self.__group_cols_rows(ref=ref, axis='rows')

    def util_get_cell_coordinates(self, ref):
        """
        Get cell coordinates

        Examples:
            >>> ws.util_get_cell_coordinates(ref='A2')
            (2, 1)
            >>> ws.util_get_cell_coordinates(ref='A')
            (None, 1)
            >>> ws.util_get_cell_coordinates(ref='1')
            (1, None)
            >>> ws.util_get_cell_coordinates(ref='-1')
            (-1, None)
        """
        ref = str(ref)

        # check if row number - i.e. all numbers
        if ref.isdigit() or (len(ref) >= 2 and ref[0] == '-' and ref[1:].isdigit()):
            return (int(ref), None)

        # check if reasonable column letter - i.e. all letters and resonable number
        if len(ref) < 4 and ref.isalpha() and openpyxl.utils.cell.column_index_from_string(ref) < 1_048_577:
            return (None, openpyxl.utils.cell.column_index_from_string(ref))

        # check if valid excel cell ref
        valid_excel_cell, col_letter, row_num = is_valid_excel_cell(ref)
        if valid_excel_cell:
            return openpyxl.utils.cell.coordinate_to_tuple(ref)

        # raise wainring if invalid
        if self.excelwriter.ref_warnings:
            warnings.warn(f'Ref "{ref}" could not be found. (You can ignore these warnings by adding ref_warnings=False to ExcelWriter())')

        # return None
        return None


    def util_get_range_coordinates(self, ref, enrich_dimensions=False):
        """
        Get cell range coordinates

        Examples:
            >>> ws.util_get_range_coordinates(ref='A2:C4')
            ((2, 1), (4, 3))
            >>> ws.util_get_range_coordinates(ref='A:C')
            ((None, 1), (None, 3))
            >>> ws.util_get_range_coordinates(ref='1:3')
            ((1, None), (3, None))
            >>> ws.util_get_range_coordinates(ref='4:-1')
            ((4, None), (-1, None))

        """
        if ':' in ref:
            ref_start, ref_end = ref.split(':')
        else:
            ref_start = ref_end = ref

        coordinates_start = self.util_get_cell_coordinates(ref_start)
        coordinates_end = self.util_get_cell_coordinates(ref_end)

        # if any invaild ref
        if coordinates_start is None or coordinates_end is None:
            return None

        row_start, col_start = coordinates_start
        row_end, col_end = coordinates_end
        if self.__class__.__name__ == 'DataframeSheet':
            ((sheet_min_row, sheet_min_column), (sheet_max_row, sheet_max_column)) = self.shape_body
        else:
            sheet_min_row = self.ws.min_row
            sheet_max_row = self.ws.max_row
            sheet_min_column = self.ws.min_column
            sheet_max_column = self.ws.max_column

        # enrich_dimensions if any ref is None/undefined
        if enrich_dimensions:
            if row_start is None:
                row_start = sheet_min_row
            if row_end is None:
                row_end = sheet_max_row - 1
            if col_start is None:
                col_start = sheet_min_column
            if col_end is None:
                col_end = sheet_max_column - 1

        # make sure lower bound at start always - so range() works
        if row_start is not None and row_end is not None and row_end < row_start:
            row_end, row_start = row_start, row_end
        if col_start is not None and col_end is not None and col_end < col_start:
            col_end, col_start = col_start, col_end

        return ((row_start, col_start), (row_end, col_end))

    def util_cell_ref_from_coordinates(self, ref):
        """Tranform (row_num, col_num) into excel ref 'A1' """
        if ref is None:
            return None

        col_letter = '' if ref[1] is None else openpyxl.utils.cell.get_column_letter(ref[1])
        row_number = '' if ref[0] is None else ref[0]
        return f'{col_letter}{row_number}'

    def util_range_ref_from_coordinates(self, ref):
        """Tranform ((start_row_num, start_col_num), (end_row_num, end_col_num)) into excel ref 'A1:C5'"""
        if ref is None:
            return None

        start_ref, end_ref = ref
        start_ref = self.util_cell_ref_from_coordinates(start_ref)
        end_ref = self.util_cell_ref_from_coordinates(end_ref)
        return f'{start_ref}:{end_ref}'


class DataframeSheet(Sheet):
    """
    DataFrame Excel Sheet class containing all logic specific to dataframe exports

    Note: All methods/function of the class Sheet also work for this class DataframeSheet
    """

    def __init__(
        self,
        excelwriter,
        df,
        sheet_name,
        startrow=0,
        startcol=0,
        index=False,
        header=True,
        style={},
        use_theme_style=True,
        col_widths={},
        col_autofit=True,
        auto_number_formatting=True,
    ):
        super().__init__(excelwriter, sheet_name, use_theme_style, col_widths)
        self.startrow = startrow
        self.startcol = startcol
        self.df = df
        self.index = df.index
        self.header = df.columns
        self.has_index = index
        self.has_header = header
        self.index_depth = index_depth = self.index.nlevels if self.has_index else 0
        self.header_depth = header_depth = self.header.nlevels if self.has_header else 0
        self.table_width = table_width = len(self.header)
        self.table_height = table_height = len(self.index)
        self.col_autofit = col_autofit
        self.auto_number_formatting = auto_number_formatting
        self.shape = ((startrow + 1, startcol + 1), (startrow + header_depth + table_height + 1, startcol + index_depth + table_width + 1))
        self.shape_header = ((startrow + 1, startcol + 1), (startrow + header_depth + 1, startcol + index_depth + table_width + 1))
        self.shape_index = ((startrow + header_depth + 1, startcol + 1), (startrow + header_depth + table_height + 1, startcol + index_depth + 1))
        self.shape_body = ((startrow + header_depth + 1, startcol + index_depth + 1), (startrow + header_depth + table_height + 1, startcol + index_depth + table_width + 1))

        # export df to excel
        self.df.to_excel(
            self.writer, sheet_name=sheet_name, startrow=startrow, startcol=startcol, index=index, header=header
        )
        self.ws = self.writer.book[sheet_name]

        # generate final styling that will apply to the dataframe export
        if use_theme_style:
            # add table style from style template
            if "table" in self.excelwriter.theme:
                for level, level_styling in self.excelwriter.theme["table"].items():
                    dict_extend_with_dict(
                        dict_obj=self.style_base, key=level, value_dict=self._extend_style_args(level_styling)
                    )

        # add number formatting
        if auto_number_formatting:
            # get all date columns
            for col_name, col_series in df.select_dtypes(include=["datetime", "datetimetz"]).items():
                self.style_base[col_name] = self._extend_style_args("date_fmt_iso")

            # get all numeric columns
            for col_name, col_series in df.select_dtypes(include=["number"]).items():
                col_series_mod = col_series.replace(0, np.nan).abs()
                low = col_series_mod.quantile(0.2)
                high = col_series_mod.quantile(0.8)

                # check if percentages
                if high < 2:
                    self.style_base[col_name] = self._extend_style_args("num_fmt_pct")
                # check if small number
                elif high < 1_000 and "int" not in str(col_series.dtype):
                    self.style_base[col_name] = self._extend_style_args("num_fmt_decimal")
                # check if large number in millions
                elif low > 10_000_000:
                    self.style_base[col_name] = self._extend_style_args("num_fmt_mm")
                # else normal number format
                else:
                    self.style_base[col_name] = self._extend_style_args("num_fmt_general")

        # add styling defined in this function
        self.style_custom = {}
        for ref, ref_style in style.items():
            dict_extend_with_dict(dict_obj=self.style_custom, key=ref, value_dict=self._extend_style_args(ref_style))

        # actually apply the final themes
        self._apply_table_style()

        # apply column widths
        _col_widths = {}
        for col_ref, col_width in col_widths.items():
            col_coordinates = self.util_get_range_coordinates(col_ref)

            if (
                col_coordinates[0] is not None
                and col_coordinates[0][1] is not None
                and col_coordinates[1] is not None
                and col_coordinates[1][1] is not None
            ):
                for col_idx in range(col_coordinates[0][1], col_coordinates[1][1] + 1):
                    _col_widths[col_idx] = col_width

        # autofit columns
        if col_autofit:
            CHARACTER_FACTOR = 1.28

            # enumerate though indeices and columns
            for i, col in enumerate(
                ([df.index.get_level_values(i) for i in range(df.index.nlevels)] if index else [])
                + [i for _, i in df.items()],
                start=startcol,
            ):
                # check if no manual col width
                if i not in _col_widths:
                    # is datetime column
                    if pd.api.types.is_datetime64_any_dtype(col):
                        _col_widths[i] = int(10 * CHARACTER_FACTOR)
                    # is string column
                    elif pd.api.types.is_string_dtype(col):
                        try:
                            length = np.quantile(col.str.len().values, 0.8)
                            _col_widths[i] = int(max(length, 4) * CHARACTER_FACTOR)
                        except:
                            # ignore error
                            pass
                    # is numeric column
                    elif pd.api.types.is_numeric_dtype(col):
                        length = len("{:,}".format(int(col.quantile(0.8))))
                        _col_widths[i] = int(max(length, 6) * CHARACTER_FACTOR)

        self.change_col_widths(_col_widths)

    def _apply_table_style(self):
        """This internal function applies the table cell styling for the to_excel() function"""
        ref_warnings = self.excelwriter.ref_warnings

        style_base = self.style_base.copy()
        style_custom = self.style_custom.copy()

        style_special_base = {**style_base.pop("base", {}), **style_custom.pop("base", {})}

        style_special_head = {**style_special_base, **style_base.pop("head", {}), **style_custom.pop("head", {})}
        style_special_index = {**style_special_base, **style_base.pop("index", {}), **style_custom.pop("index", {})}
        style_special_body = {**style_special_base, **style_base.pop("body", {}), **style_custom.pop("body", {})}

        style_non_special = {}
        for iter_dict in [style_base, style_custom]:
            for ref, ref_style in iter_dict.items():
                style_coordinates = self.util_get_range_coordinates(ref)

                # check if invalid cell ref
                if ref_warnings and (style_coordinates[0] is None or style_coordinates[1] is None):
                    warnings.warn(
                        f'Styling ref "{ref}" could not be found. (You can turn off these warnings by adding ref_warnings=False to ExcelWriter())'
                    )

                # all good - valid cell style range
                else:
                    dict_extend_with_dict(
                        dict_obj=style_non_special, key=style_coordinates, value_dict=self._extend_style_args(ref_style)
                    )

        # Apply table heading styling
        if self.has_header:
            ((start_row, start_col), (end_row, end_col)) = self.shape_header
            for row_num in range(start_row, end_row):
                for col_num in range(start_col, end_col):
                    cell_style = get_custom_styles(
                        cell_row_num=row_num,
                        cell_col_num=col_num,
                        styles_custom=style_non_special,
                        ignore_entire_rows_or_cols=True,
                    )
                    # print('Header', row_num, col_num)
                    self.apply_cell_style(row_num=row_num, col_num=col_num, style={**style_special_head, **cell_style})

        # Apply table index styling
        if self.has_index:
            ((start_row, start_col), (end_row, end_col)) = self.shape_index
            for row_num in range(start_row, end_row):
                for col_num in range(start_col, end_col):
                    cell_style = get_custom_styles(
                        cell_row_num=row_num,
                        cell_col_num=col_num,
                        styles_custom=style_non_special,
                        ignore_entire_rows_or_cols=True,
                    )
                    # print('Index', row_num, col_num)
                    self.apply_cell_style(row_num=row_num, col_num=col_num, style={**style_special_index, **cell_style})

        # Apply table body styling
        ((start_row, start_col), (end_row, end_col)) = self.shape_body
        for row_num in range(start_row, end_row):
            for col_num in range(start_col, end_col):
                cell_style = get_custom_styles(
                    cell_row_num=row_num,
                    cell_col_num=col_num,
                    styles_custom=style_non_special,
                    ignore_entire_rows_or_cols=False,
                )
                # print('Body', row_num, col_num)
                self.apply_cell_style(row_num=row_num, col_num=col_num, style={**style_special_body, **cell_style})


    def add_data_validation(self,
            ref: Union[str, List[str]],
            type: str, #Literal["list", "whole", "decimal", "date", "time", "textLength", "formula"],
            props: Union[str, List[str], int, float, Tuple[int, float]] = None,
            operator: str = None, #Literal["between", "greaterThan", "greaterThanOrEqual", "equal", "notEqual", "lessThan", "lessThanOrEqual"] = None,
            **kwargs
        ):
        """
        Add data validation to worksheet

        Args:
            ref (str or list of str): Cell range reference e.g. "A1:C5" or "employees" or ["A1:C5", "A21:Z55", "employees:RoE"]
            type (str): Data valiation type. Options: list, whole, decimal, date, time, textLength, formula
            operator (str): Validation mathematical operator: "between", "greaterThan", "greaterThanOrEqual", "equal", "notEqual", "lessThan", "lessThanOrEqual"
            props (any): Depending on vaildation type the properties: list = ["Option 1", "Option 2"], numeric = single number OR upper & lower boundary (1, 100), formla
            kwargs (dict): Any futher kwarg specified in [openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.datavalidation.html](https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.datavalidation.html#openpyxl.worksheet.datavalidation.DataValidation)

        Examples:
            - ```sheet1.add_data_validation(ref="employees", type="whole")```
            - ```sheet1.add_data_validation(ref="A1:C5", type="whole")```
            - ```sheet1.add_data_validation(ref="RoE", type="decimal", operator="greaterThan", props=0)```
            - ```sheet1.add_data_validation(ref="A1:C5", type="textLength", props=10)```
            - ```sheet1.add_data_validation(ref="A1:C5", type="list", props=["Yes", "No"])```
            - ```sheet1.add_data_validation(ref="A1:C5", type="whole", operator="between", props=[0, 100])```
        """
        if isinstance(ref, str):
            ref = [ref]
        ref = [self.util_range_ref_from_coordinates(self.util_get_range_coordinates(i, enrich_dimensions=True)) if i in self.header else i for i in ref]
        super().add_data_validation(ref=ref, type=type, props=props, operator=operator, **kwargs)


    def util_get_cell_coordinates(self, ref):
        """
        Get cell coordinates

        Examples:
            >>> ws.util_get_cell_coordinates(ref='employees')
            (N, 3)
            >>> ws.util_get_cell_coordinates(ref='A2')
            (2, 1)
            >>> ws.util_get_cell_coordinates(ref='A')
            (None, 1)
            >>> ws.util_get_cell_coordinates(ref='1')
            (1, None)
            >>> ws.util_get_cell_coordinates(ref='-1')
            (-1, None)
        """
        ref = str(ref)
        if not ref.isdigit() or not (len(ref) >= 2 and ref[0] == '-' and ref[1:].isdigit()):
            if ref in self.header:
                ref = openpyxl.utils.cell.get_column_letter((self.header.get_loc(ref) + 1) + self.index_depth + self.startcol)
        return super().util_get_cell_coordinates(ref)


class ExcelWriter:
    """
    Class for writing DataFrame objects into excel sheets.

    Example:
        Output pandas dataframe quickly with beautiful formatting.

        ```python
        from beautifulexcel import ExcelWriter

        with ExcelWriter('workbook.xlsx', mode='r', style='elegant_blue') as writer:
            ws1 = writer.to_excel(df1, sheetname='My Sheet', mode='a', startrow=0, startcol=0)
            ws2 = writer.to_excel(df2, sheetname='My Sheet', mode='a', startrow=20, startcol=0)
        ```
    """

    def __init__(
        self,
        file: str,
        mode: str = "replace", #Literal["replace", "modify"] = "replace",
        if_sheet_exists: str = "replace", #Literal["error", "new", "replace", "overlay"] = "replace",
        theme: str = "elegant_blue",
        ref_warnings: bool = True,
        date_format: str = None,
        datetime_format: str = None,
        engine_kwargs: Any = {},
        **kwargs,
    ):
        """
        Args:
            file (str): Path to xls or xlsx or ods file
            mode (str): If the file already exists you can either "replace" or "modify" it
            if_sheet_exists (str): If a excel sheet already exists raise an "error", create a "new" sheet with a different name, "replace" the existing sheet with the new one, or "overlay" the new contents with the old ones
            theme (str): Excel style name or path to theme yaml file
            date_format (str): Format string for dates written into Excel files (e. g. 'YYYY-MM-DD')
            datetime_format (str): Format string for datetime objects written into Excel files. (e. g. 'YYYY-MM-DD HH:MM:SS')
            engine_kwargs (str): keywords passed though to openpyxl in "replace"-mode: openpyxl.Workbook(**engine_kwargs); "modify"-mode: openpyxl.load_workbook(file, **engine_kwargs)

        Example:
            ```python
            from beautifulexcel import ExcelWriter

            with ExcelWriter('workbook.xlsx', mode='r', style='elegant_blue') as writer:
                ...
            ```
        """
        self.file = file
        self.mode = mode
        self.if_sheet_exists = if_sheet_exists
        self.theme_name = theme
        self.ref_warnings = ref_warnings
        self.date_format = date_format
        self.datetime_format = datetime_format
        self.engine_kwargs = engine_kwargs

        # modify existing file
        if os.path.isfile(file) and mode == "modify":
            self.writer = pd.ExcelWriter(
                file,
                engine="openpyxl",
                mode="a",
                if_sheet_exists=if_sheet_exists,
                date_format=date_format,
                datetime_format=datetime_format,
                engine_kwargs=engine_kwargs,
                **kwargs,
            )
            # self.writer.book = openpyxl.load_workbook(file, **engine_kwargs)
            # self.writer.sheets = {ws.title: ws for ws in self.writer.book.worksheets}
            self.file_mode = "modify"

        # create new file
        else:
            self.writer = pd.ExcelWriter(
                file,
                engine="openpyxl",
                mode="w",
                if_sheet_exists=None,
                date_format=date_format,
                datetime_format=datetime_format,
                engine_kwargs=engine_kwargs,
                **kwargs,
            )
            # self.writer.book = openpyxl.Workbook(**engine_kwargs)
            self.file_mode = "replace"

        # explicitly no theme defined
        if theme is None or len(theme) == 0:
            self.theme = {}

        # theme defined
        else:
            # standard theme from package
            if "." not in theme:
                this_file_path = os.path.dirname(os.path.abspath(__file__))
                theme = os.path.join(this_file_path, "themes", f"{theme}.yml")

            # read in user yml theme file
            try:
                with open(theme, "r") as file:
                    theme = yaml.safe_load(file)
                    theme_converted = {}
                    # flatten after level 2
                    for level1_name, level1 in theme.items():
                        if level1_name not in theme_converted:
                            theme_converted[level1_name] = {}
                        for level2_name, level2 in level1.items():
                            if level2_name not in theme_converted:
                                theme_converted[level1_name][level2_name] = {}
                            theme_converted[level1_name][level2_name] = flatten_dict(level2)
                    self.theme = theme_converted
            except yaml.YAMLError as exc:
                raise Exception(f"Error when reading in theme file from path '{theme}':", exc)

    def __enter__(self):
        if not hasattr(self, "file"):
            raise Exception(
                "Wrong usage! Please run it this way:\n>>> from beautifulexcel import ExcelWriter\n>>> with ExcelWriter('workbook.xlsx', mode='r', theme='elegant_blue') as writer:\n>>>     ws1 = writer.write_df(df1, sheetname='My Sheet'"
            )
        return self

    def save(self):
        if callable(getattr(self.writer, "save", None)):
            self.writer.save()
        else:
            self.writer._save()

    def __exit__(self, type, value, traceback):
        # ToDo: add exception handling here
        self.save()
        self.writer = None
        self = None

    def to_excel(
        self,
        df: pd.DataFrame,
        sheet_name: str,
        startrow: int = 0,
        startcol: int = 0,
        index: bool = False,
        header: bool = True,
        style: dict = {},
        use_base_style: bool = True,
        col_widths: dict = {},
        col_autofit: bool = True,
        auto_number_formatting: bool = True,
    ) -> DataframeSheet:
        """
        Export pandas Datafame to excel.

        Args:
            df (pd.DataFrame): Pandas Dataframe to export
            sheet_name (str): Sheet name
            startrow (int): Upper left cell row to dump dataframe (zero indexed)
            startcol (int): Upper left cell column to dump data rame (zero indexed)
            index (bool): Write row names/index as first column/s
            header (bool): Write column names/header as first row/s
            style (dict): Style dictionary with key referencing the cell/column/row and value the style to apply e.g. {'RoE': 'bg_light_blue', 'D:E': {'fill': 'FFEEB7'}}
            use_base_style (bool): Apply the excel workbook "theme" set in ExcelWriter()
            col_widths (dict): Define column widths manually with key referencing the column and value the width e.g. {'A:C': 20, 'F': 10, 'employees': 40}
            col_autofit (bool): Automatically change column width to fit content best
            auto_number_formatting (bool): Automatically detect number format and change excel format

        Returns:
            beautifulexcel.DataframeSheet

        Example:
            ```python
            from beautifulexcel import ExcelWriter

            with ExcelWriter('workbook.xlsx', mode='r', style='elegant_blue') as writer:
                ws1 = writer.to_excel(df1, sheetname='My Sheet', mode='a', startrow=0, startcol=0)
            ```
        """
        df_sheet = DataframeSheet(
            excelwriter=self,
            df=df,
            sheet_name=sheet_name,
            startrow=startrow,
            startcol=startcol,
            index=index,
            header=header,
            style=style,
            use_theme_style=use_base_style,
            col_widths=col_widths,
            col_autofit=col_autofit,
            auto_number_formatting=auto_number_formatting,
        )

        return df_sheet


if __name__ == "__main__":
    example_df = pd.DataFrame(
        {
            "client": ["A", "B", "C", "D", "E", "F", "G", "H"],
            "industry": [
                "ASEET MANAGEMENT",
                "ASEET MANAGEMENT",
                "BANK",
                "INSURANCE",
                "VERY VERY VERY VERY VERY LONG INSUSTRY NAME",
                "BANK",
                "COMMODITY BROKER",
                "INSURANCE",
            ],
            "employees": [25_000, 17_000_000, 14, 9_000, 12_000_000, 9_000, np.nan, 4_000_000],
            "inception": [
                datetime.datetime(2022, 1, 1),
                datetime.datetime(2000, 10, 30),
                datetime.datetime(2010, 5, 6),
                np.nan,
                datetime.datetime(1997, 1, 1),
                datetime.datetime(1962, 1, 1),
                datetime.datetime(2003, 11, 10),
                datetime.datetime(2022, 1, 1),
            ],
            "last_contact": [
                datetime.datetime(2022, 1, 1, 10, 2, 5),
                datetime.datetime(2000, 10, 30),
                datetime.datetime(2010, 5, 6),
                np.nan,
                datetime.datetime(1997, 1, 1),
                datetime.datetime(1962, 1, 1),
                datetime.datetime(2003, 11, 10, 10, 2, 5),
                datetime.datetime(2022, 1, 1),
            ],
            "RoE": [0.05, -0.05, 0.15, 1.05, -0.02, np.nan, 0.08, 0.05],
            "revenue": [
                50_000_000.4387,
                np.nan,
                63_000_000.4387,
                25_000.4387,
                25_000_000.4387,
                -50_000_000.4387,
                76_000_000.4387,
                25_000.4387,
            ],
        }
    )
    example_df.set_index("client", inplace=True)

    with ExcelWriter("workbook.xlsx", mode="r", theme="elegant_blue") as writer:
        ws1 = writer.to_excel(
            example_df,
            style={"RoE": "bg_light_blue", "D:E": {"fill": "FFEEB7"}},
            sheet_name="My Sheet",
            startrow=0,
            startcol=0,
            index=True,
            col_widths={"employees": 100},
        )
        ws1.add_data_validation(ref="employees", type="list", props=["Y", "N"])
        ws1.group_columns("revenue:B")
        ws1.merge_cells("B8:C9")

    # example_df.to_excel('test.xlsx', sheet_name='My Sheet', startrow=0, startcol=0, index=True)
